import csv
import io
import json
from typing import Literal

from .schemas import (
    ProtocolType,
    SubjectCreateRequest,
)

BATCH_LIMIT = 500
VALID_SEX_VALUES = {"female", "male", "unknown"}
VALID_PROTOCOLS = {"static_posture", "adams_forward_bend", "squat"}


def detect_format(filename: str) -> Literal["csv", "excel", "json"]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return "csv"
    if lower.endswith((".xlsx", ".xls")):
        return "excel"
    if lower.endswith(".json"):
        return "json"
    raise ValueError(f"Unsupported file format: {filename}")


def _decode_bytes(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb2312", "latin-1"):
        try:
            return content.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            continue
    return content.decode("utf-8", errors="replace")


def parse_csv(content: bytes) -> list[dict]:
    text = _decode_bytes(content)
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise ValueError("CSV file has no header row")
    rows = []
    for row in reader:
        cleaned = {k.strip(): v.strip() if v else "" for k, v in row.items()}
        if any(cleaned.values()):
            rows.append(cleaned)
    return rows


def parse_excel(content: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel import. Install it with: pip install openpyxl")

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no active sheet")

    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h else "" for h in next(rows_iter, [])]
    if not headers or all(h == "" for h in headers):
        raise ValueError("Excel file has no header row")

    rows = []
    for row in rows_iter:
        values = [str(v).strip() if v is not None else "" for v in row]
        if not any(values):
            continue
        padding = [""] * max(0, len(headers) - len(values))
        record = dict(zip(headers, values + padding))
        rows.append(record)

    wb.close()
    return rows


def parse_json(content: bytes) -> list[dict]:
    text = _decode_bytes(content)
    data = json.loads(text)
    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list):
        raise ValueError("JSON content must be an array of objects or a single object")
    return data


def parse_file(filename: str, content: bytes) -> list[dict]:
    fmt = detect_format(filename)
    if fmt == "csv":
        return parse_csv(content)
    if fmt == "excel":
        return parse_excel(content)
    return parse_json(content)


def validate_subject_row(row: dict, row_index: int) -> tuple[SubjectCreateRequest | None, list[str]]:
    errors: list[str] = []
    display_name = row.get("display_name", "").strip()
    if not display_name:
        errors.append("缺少必填字段: display_name")
    elif len(display_name) > 80:
        errors.append(f"display_name 超过80字符上限: {len(display_name)}")

    sex_raw = row.get("sex", "").strip().lower()
    sex: Literal["female", "male", "unknown"] = "unknown"
    if sex_raw:
        if sex_raw in VALID_SEX_VALUES:
            sex = sex_raw  # type: ignore[assignment]
        else:
            errors.append(f"无效的 sex 值 '{sex_raw}'，可选: female, male, unknown")

    age_raw = row.get("age", "").strip()
    age: int | None = None
    if age_raw:
        try:
            age_val = int(age_raw)
            if age_val < 3 or age_val > 120:
                errors.append(f"age 超出有效范围 (3-120): {age_val}")
            else:
                age = age_val
        except ValueError:
            errors.append(f"age 不是有效整数: '{age_raw}'")

    height_raw = row.get("height_cm", "").strip()
    height_cm: float | None = None
    if height_raw:
        try:
            h = float(height_raw)
            if h < 60 or h > 240:
                errors.append(f"height_cm 超出有效范围 (60-240): {h}")
            else:
                height_cm = h
        except ValueError:
            errors.append(f"height_cm 不是有效数字: '{height_raw}'")

    notes = row.get("notes", "").strip()
    if len(notes) > 500:
        errors.append(f"notes 超过500字符上限: {len(notes)}")

    if errors:
        return None, errors

    return SubjectCreateRequest(
        display_name=display_name,
        sex=sex,
        age=age,
        height_cm=height_cm,
        notes=notes,
    ), []


def validate_session_row(row: dict, row_index: int) -> tuple[dict | None, list[str]]:
    """Validate a screening session import row.

    Returns (validated_data | None, errors).
    validated_data has keys: subject_id, subject_display_name, protocols
    """
    errors: list[str] = []
    subject_id = row.get("subject_id", "").strip() or None
    subject_name = row.get("subject_display_name", "").strip() or None

    if not subject_id and not subject_name:
        errors.append("缺少必填字段: subject_id 或 subject_display_name (至少提供一个)")
    if subject_id and subject_name:
        errors.append("subject_id 和 subject_display_name 只能提供一个")

    protocols_raw = row.get("protocols", "").strip()
    protocols: list[ProtocolType] = []
    if protocols_raw:
        for p in protocols_raw.split(","):
            p = p.strip()
            if p in VALID_PROTOCOLS:
                protocols.append(p)  # type: ignore[arg-type]
            else:
                errors.append(f"无效的协议类型 '{p}'，可选: static_posture, adams_forward_bend, squat")
    else:
        protocols = ["static_posture", "adams_forward_bend", "squat"]

    if errors:
        return None, errors

    return {
        "subject_id": subject_id,
        "subject_display_name": subject_name,
        "protocols": protocols,
    }, []
